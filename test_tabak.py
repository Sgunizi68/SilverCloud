#!/usr/bin/env python3
"""Test script to inspect the Excel file structure"""
import sys

try:
    import openpyxl
    wb = openpyxl.load_workbook(r'C:\Users\sg.msa\Downloads\Şube Bazlı Toplam Tabak Sayısı - 1.02.2026 06_00_00 - 1.03.2026 06_00_00.xlsx', read_only=True)
    print('Sheet names:', wb.sheetnames)
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f'\n--- Sheet: {sname} ---')
        count = 0
        for i, row in enumerate(ws.rows):
            if i <= 4:  # First 5 rows
                vals = [cell.value for cell in row]
                types = [type(cell.value).__name__ for cell in row]
                print(f'  Row {i+1} types={types}: {vals}')
            count += 1
        print(f'  Total rows iterated: {count}')
    wb.close()
except Exception as e:
    print(f'openpyxl error: {e}')
    import traceback; traceback.print_exc()
